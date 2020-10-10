# -*- coding: utf-8 -*-
"""
Created on Sat Oct 10 16:22:55 2020

@author: heart
"""

#import pandas
import openpyxl
#from openpyxl import Workbook
from openpyxl.styles import Alignment

def compare_excel_2_3(infile1, infile2, outfile):    
    ### Read excel file
#    wb1 = openpyxl.load_workbook(f"/home/l6oi/IDDOC/Excel_File_Repository/VTAPM_2_2_F101_Annual_2019.xlsx")
#    wb2 = openpyxl.load_workbook(f"/home/l6oi/IDDOC/Excel_File_Repository/VTAPM_2_2_F101_Annual_2020.xlsx")
    wb1 = openpyxl.load_workbook(infile1)
    wb2 = openpyxl.load_workbook(infile2)
    
    ### List of all sheet names
    #wb1.sheetnames
    #wb2.sheetnames
    
    ### Get sheet by name
    #ws = wb['F101']
    
    ### Get sheet by Index
    ws1 = wb1.worksheets[0]
    ws2 = wb2.worksheets[0]
    
    header1 = ws1['A1'].value
    header2 = ws2['A1'].value
    header1_lines = header1.split("\n")
    
    header = header2 + '\nvs.\n' +  header1_lines[3] 
    ws2['A1'] = header
        
    # {:.2%}.format(b2)
    ws1_b2 = ws1['B2'].value
    ws2_b2 = ws2['B2'].value
    
    cell_b2 = ws2_b2 + '\n' + "vs." + '\n' + ws1_b2
    ws2['B2'] = cell_b2
    
    
    ws1_d2 = ws1['D2'].value
    ws2_d2 = ws2['D2'].value
    
    cell_d2 = ws2_d2 + '\nvs.' + '\n' + ws1_d2
    ws2['D2'] = cell_d2
    
    
    ws1_b3 = ws1['B3'].value
    ws2_b3 = ws2['B3'].value
    
    cell_b3 = "{:.2%}".format((ws2_b3 - ws1_b3)/ws1_b3)
    ws2['B3'] = cell_b3
    
    
    ws1_d3 = ws1['D3'].value
    ws2_d3 = ws2['D3'].value
    
    cell_d3 = "{:.2%}".format((ws2_d3 - ws1_d3)/ws1_d3)
    ws2['D3'] = cell_d3
    
    
    has_empty_cell = False
    
    ### To get the changes year over year 
    for row in range(5,13):
        for col in range(2,6):
            v1 = ws1.cell(row, col).value
            v2 = ws2.cell(row, col).value
            
            if isinstance(v1, str):
                v1 = v1.replace(',','').replace('%','')
                
            if isinstance(v2, str):
                v2 = v2.replace(',','').replace('%','')

            if v1 is None or v2 is None:
                has_empty_cell = True
                
            if col == 2 or col == 4:
                v1 = int(v1)
                v2 = int(v2)
            else:
                v1 = float(v1)
                v2 = float(v2)
                
            change = "{:.2%}".format((v2-v1)/v1)
            ws2.cell(row, col).value = change
    
    ws2.row_dimensions[1].height = 100    
    ws2.row_dimensions[2].height = 50
    al = Alignment(horizontal='center', vertical='center')
    ws2.cell(2, 2).alignment = al

    if has_empty_cell:
        ws2['A15'] = "Note: The report contains cells with blank values"
    else:
        ws2['A15'] = "Note: The report contains no blank cells"
        
#    wb2.save("/home/l6oi/IDDOC/output/VTAPM_2_2_F101_Annual_Report_Validation_2020.xlsx")
    wb2.save(outfile)           
    
